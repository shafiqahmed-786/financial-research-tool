from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


def generate_excel(income_data, metadata, metrics, output_path):

    wb = Workbook()

    # =================================================
    # Sheet 1 — Income Statement
    # =================================================
    ws = wb.active
    ws.title = "Income Statement"

    periods = income_data["periods"]
    rows = income_data["rows"]

    ws["A1"] = "Income Statement"
    ws["A1"].font = Font(size=14, bold=True)

    ws["A2"] = f"Statement Type: {metadata.get('statement_type')}"
    ws["A3"] = f"Currency: {metadata.get('currency')}"

    header_row = 5

    ws.cell(row=header_row, column=1).value = "Line Item"
    ws.cell(row=header_row, column=1).font = Font(bold=True)

    for i, period in enumerate(periods):
        cell = ws.cell(row=header_row, column=i + 2)
        cell.value = period
        cell.font = Font(bold=True)

    for r_index, row in enumerate(rows):
        ws.cell(row=r_index + header_row + 1, column=1).value = row["line_item"]

        for c_index, period in enumerate(periods):
            value = row.get(period)
            cell = ws.cell(
                row=r_index + header_row + 1,
                column=c_index + 2
            )
            cell.value = value
            cell.number_format = '#,##0.00'

    ws.freeze_panes = "B6"

    for col in range(1, len(periods) + 2):
        ws.column_dimensions[get_column_letter(col)].width = 18

    # =================================================
    # Sheet 2 — Financial Metrics
    # =================================================
    ws2 = wb.create_sheet("Financial Metrics")

    ws2["A1"] = "Financial Metrics"
    ws2["A1"].font = Font(size=14, bold=True)

    row_cursor = 3

    green_fill = PatternFill(start_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", fill_type="solid")

    for metric_name, values in metrics.items():

        title = metric_name.replace("_", " ").title()

        ws2.cell(row=row_cursor, column=1).value = title
        ws2.cell(row=row_cursor, column=1).font = Font(bold=True)

        # Period headers
        col_cursor = 2
        for period in values.keys():
            ws2.cell(row=row_cursor, column=col_cursor).value = period
            ws2.cell(row=row_cursor, column=col_cursor).font = Font(bold=True)
            col_cursor += 1

        row_cursor += 1
        col_cursor = 2

        for value in values.values():

            cell = ws2.cell(row=row_cursor, column=col_cursor)
            cell.value = value

            if value is not None:
                cell.number_format = '0.00%'
                if value > 0:
                    cell.fill = green_fill
                else:
                    cell.fill = red_fill

            col_cursor += 1

        row_cursor += 3

    # =================================================
    # Sheet 3 — Executive Summary
    # =================================================
    ws3 = wb.create_sheet("Executive Summary")

    ws3["A1"] = "Executive Summary"
    ws3["A1"].font = Font(size=14, bold=True)

    latest_period = income_data["periods"][0]

    ws3["A3"] = f"Latest Period: {latest_period}"

    if "revenue_growth" in metrics:
        growth = metrics["revenue_growth"].get(latest_period)
        if growth:
            ws3["A5"] = f"Revenue grew by {round(growth, 2)}% in {latest_period}"

    if "pat_margin" in metrics:
        margin = metrics["pat_margin"].get(latest_period)
        if margin:
            ws3["A6"] = f"PAT Margin in {latest_period}: {round(margin, 2)}%"

    wb.save(output_path)
